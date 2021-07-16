# -*- coding: utf-8 -*-
import os
from datetime import datetime
import sqlite3
import pandas as pd
import openpyxl as op


# Получение имени таблицы для дальнейшего взаимодействия
def get_table(path="./second_input"):
    print("Enter table name...")
    name = str(input())
    print("table name is " + name)
    os.chdir(path)
    names_list = os.listdir()
    if not(name in names_list):
        print("name not found.")
        pass
    else:
        print("name found in folder.")
        return name


# Создание большого dataframe для парсинга
def pars_table(name):
    name_of_table = name
    list_of_headers = ["Номенклатура", "Артикул", "Характеристики", "Единицы измерения", "Оптовая", "Розничные цены СПБ", "Остаток", "Заказать"]

    dictionary_to_write = { header: 0 for header in list_of_headers}
    dictionary_to_write["name"] = []

    tamplate_of_data_frame = {header: [] for header in list_of_headers}


    frame_of_tables = pd.DataFrame(tamplate_of_data_frame)

    work_book = op.load_workbook(name)
    # Номер заказа|Статус|Время создания заказа|Время оплаты|Стоимость товаров, Руб|Стоимость доставки, Руб|Сумма заказа, Руб|Скидка магазина, Руб|Оплачено клиентом, Руб|Сумма возврата, Руб|Возвраты|Товары|Артикул|Id товаров|Примечания к заказу (покупателя)|Примечания к заказу (продавца)|Имя получателя|Страна|Штат/провинция|Город|Адрес|Индекс|Телефон|Способ доставки|Отгрузка истекает|Трекинг номер|Время отправки|Время подтверждения покупателем|
    for name in work_book.sheetnames:
        sheet = work_book[name]
        maxrow = sheet.max_row  # Получение максимального значения строк данного листа книги
        maxcol = sheet.max_column  # Получение максимального столбцов строк данного листа книги
        print("rows: cols:")
        print(maxrow, maxcol)
        print("Generate dataFrame...")
        # Номер заказа|Статус|Время создания заказа|Время оплаты|Стоимость товаров, Руб|Стоимость доставки, Руб|Сумма заказа, Руб|Скидка магазина, Руб|Оплачено клиентом, Руб|Сумма возврата, Руб|Возвраты|Товары|Артикул|Id товаров|Примечания к заказу (покупателя)|Примечания к заказу (продавца)|Имя получателя|Страна|Штат/провинция|Город|Адрес|Индекс|Телефон|Способ доставки|Отгрузка истекает|Трекинг номер|Время отправки|Время подтверждения покупателем|
        for row in sheet.iter_rows(min_row=3, min_col=1, max_row=maxrow, max_col=maxcol):
            index_of_headers = 0
            for cell in row:
                if index_of_headers == 0:
                    index = index_of_headers
                    if cell.value == "":
                        dictionary_to_write[list_of_headers[index]] = "None"
                    else:
                        dictionary_to_write[list_of_headers[index]] = cell.value

                if index_of_headers > 11:
                    index = index_of_headers - 11
                    #print(index)
                    # if cell.value != None:
                    if cell.value == "" or cell.value == None:
                        dictionary_to_write[list_of_headers[index]] = "None"
                    else:
                        dictionary_to_write[list_of_headers[index]] = cell.value

                index_of_headers += 1
            dictionary_to_write["name"] = name_of_table
            frame_of_tables = frame_of_tables.append(dictionary_to_write, ignore_index=True)
    #set_option для выведения всех колонок ataFrame
    pd.set_option('display.max_columns', None)
    #print(frame_of_tables)
    return frame_of_tables


# запись выбранных  значений в sql бд
def make_table_for_sql(frame_of_tables, conn):
    first_grade_list = ["Оптика", "Аксесуары для оптика", "Горны и рожки", "Дартс", "Засидки и лабазы", "Звуковые имитаторы",
                       "Кронштейны и инструменты", "Манки, приманки, нейтрализаторы", "Мишени", "Ножи", "Ножи",
                       "Одежда и обувь", "Оптоволоконные мушки", "Пневматика", "Рогатки", "Снаряжение", "Спортивная стрельба",
                       "Средства для чистки и смазки оружия", "Сумки и рюкзаки", "Термосы", "Товары для собак",
                       "Тюнинг оружия", "Фонарики и ЛЦУ", "Холодная пристрелка оружия", "Чехлы и кейсы для оружия", "Чучела", "Уцененные товары"]

    second_grade_list = ["Оптические прицелы", "Бинокли","Дальномеры", "Зрительные трубы", "Коллиматорные прицелы",
                         "Монокуляры", "Тепловизионные приборы", "Цифровые приборы День/ночь",'Крышки для прицелов "Butler Creek"',
                         "Разное", "Ameristep (США)", "ShotTime", "Аксессуары", "iHunt", 'Звуковые имитаторы "Cass Creek" (США)',
                         'Звуковые имитаторы "Mini Phantom" (США)', 'Звуковые имитаторы "MUNDI SOUND" (Испания)', 'Инструменты',
                         'Кронштейны', 'Манки Buck Expert (Канада)', "Манки FAULK'S (США)","Манки Helen Baud (Франция)",
                         "Манки Hubertus (Германия)", "Манки Mankoff (Россия)", "Манки Mick Lacy (Канада)",
                         "Манки Nordik Predator (Швеция)", "Манки PRIMOS", "Нейтрализаторы запаха", "Приманки Buck Expert (Канада)",
                         "LionSteel(Италия)", "McNETT TACTICAL(США)", "Morakniv(Швеция)", "Opinel(Франция)", "Sanrenmu(Китай)",
                         "Tekut(Китай)", "Маски", "Перчатки", "Стрелковые и разгрузочные жилеты", "HIVIZ (США)", "NIMAR (Италия)",
                         "TRUGLO (США)", "БАЛЛОНЧИКИ CO2", "Пневматические пистолеты", "Пульки и шарики для пневматики",
                         "СТРАЙКБОЛ", "Рогатки", "Шарики для рогаток", "Антабки", "Камуфляжная лента", "Кейсы и ящики для патронов, снаряжения и чистки оружия",
                         "Кобуры и сумки для пистолетов", "Ремни, патронташи и подсумки", "Сошки и опоры для оружия",
                         "Стулья", "Тыльники", "Машинки для метания", "Стрелковые наушники и беруши","Стрелковые очки",
                         "Масла и смазки", "Наборы для чистки", "Рюкзаки", "Сумки", "Ягдташи", "THERMOcafe (гарантия 1 год)",
                         "THERMOS (гарантия 5 лет)", "Термосумки и хладоэлементы", "ATA MOULD (Турция)", "Pufgun (Россия)",
                         "Red Heat/AKADEMIA (Россия)", "МОЛОТ (Россия)", "Тюнинг Hartman (Россия)", "Тюнинг Leapers UTG (США)",
                         "Тюнинг VS (Россия)", "Тюнинг Россия", "Лазерные целеуказатели", "Разное", "Фонари LEAPERS UTG (США)",
                         "Фонари NexTORCH (Китай)", "Фонари Sightmark", "Firefield (США)", "Nikko Stirling", "Red-I (ЮАР)",
                         "ShotTime", "Sightmark (США)", "Замки на оружие", "Кейсы для оружия", "Чехлы", "Аксессуары",
                         "Чучела BIRDLAND (Китай)", "Чучела SPORTPLAST (Италия)", "Оптика - УТ", "Тепловизионные приборы -УТ"]

    third_grade_list = ["LEAPERS UTG (США)", "LEUPOLD (США)", "NIKKO STIRLING", "SWAROVSKI (Австрия)", "TARGET OPTIC (Китай)",
                        "Прицелы (Китай)", "Бинокли BUSHNELL (США)", "Бинокли GAUT (Китай)", "Бинокли LEUPOLD (США)",
                        "Бинокли NIKON (Япония)", "Бинокли STEINER (Германия)", "Бинокли VANGUARD (Китай)",
                        "GAUT", "Leupold / Redfield", "Nikko Stirling", "Nikon", "Sightmark", "Vortex", "Коллиматорные прицелы Aimpoint (Швеция)",
                        "Коллиматорные прицелы Firefield (США)", "Коллиматорные прицелы Holosun (США)", "Коллиматорные прицелы Leapers UTG (США)",
                        "Коллиматорные прицелы Redring (Швеция)", "Коллиматорные прицелы SIGHTMARK (США)",
                        "Коллиматорные прицелы Target Optic (Китай)", "Коллиматорные прицелы Tokyo Scope/Hakko (Япония)",
                        "Тепловизионные монокуляры", "Тепловизионные прицелы", "Аксессуары", "Прицелы", "ATN", "Contessa Alessandro (Италия)",
                        "EAW Apel (Германия)", "Innomount (Германия)", "Leapers UTG (США)", "Leupold (США)",
                        "MAK (Германия)", "RECKNAGEL (Германия)", "Кронштейны (Китай)", "Кронштейны (Россия и Белоруссия)",
                        "Пистолеты Cybergun(Swiss Arms, Франция)", "Пистолеты Stalker", "Пульки для пневматики", "Шарики для пневматики",
                        "Allen", "McNett", "Negrini (Италия)", "PLANO (США)", "Разное",
                        "Патронташи", "Подсумки", "Ремни", "Опоры для оружия", "Сошки для оружия", "Walkstool (Швеция)",
                        "3M Peltor(США)", "Allen(США)", "Artilux(Швейцария)", "CassCreek(США)", "Howard Leight(США)",
                        "MSA(Швеция)", "Pro Ears(США)", "Rifleman(США)", "ShotTime", "Стрелковые очки 3M (США)",
                        "Стрелковые очки Allen", "Стрелковые очки ARTILUX (Швейцария)", "Стрелковые очки Randolph Engineering Inc.(США)",
                        "Стрелковые очки Stalker (Тайвань)", "BALLISTOL (Германия)", "Bore Tech (США)", "Butch's (США)",
                        "Iosso (США)", "KANO (США)", "KG Industries (США)", "Milfoam (Финляндия)", "SWEET'S (Австралия)",
                        "Waffen Care (Германия)", "Треал-М (Россия)", "A2S GUN (Россия)", "Bore Tech (США)",
                        "DAC - Универсальные наборы (США)", "HOPPE'S (США)", "J.DEWEY (США)", "NIMAR (Италия)",
                        "ShotTime", "Stil Crin (Италия)", "Лазерные целеуказатели Holosun", "Лазерные целеуказатели Leapers",
                        "Аксессуары для фонарей NexTORCH(Китай)", "Фонари NexTORCH(Китай)", "NEGRINI (Италия)",
                        "PLANO (США)", "ALLEN (США)", "LEAPERS UTG (США)", "ВЕКТОР (Россия)"]

    list_of_headers = ["Номенклатура", "Артикул",  "Оптовая", "Розничные цены СПБ", "Остаток"]
    dict_to_write = {header: list() for header in list_of_headers}
    status_to_write = {"first_status": (), "second_status": (), "third_status": ()}
    list_to_sql = []
    status_to_write["first_status"] = "None"
    status_to_write["second_status"] = "None"
    status_to_write["third_status"] = "None"
    #Остатки СПБ 21.05.2021.xlsx
    cursor = conn.cursor()
    frame_of_tables_loc = frame_of_tables[list_of_headers]
    print("Ganerate dataFrame with statuses")
    for loc in range(len(frame_of_tables_loc.index)):
        row = frame_of_tables_loc.iloc[loc]
        #print(row)
        #print(row.to_dict())
        for elem in row:
            if type(elem) == str:
                elem = elem.strip()
                if elem == "":
                    elem = "None"

            if elem in first_grade_list:
                status_to_write["first_status"] = elem
                status_to_write["second_status"] = "None"
                status_to_write["third_status"] = "None"
                break

            elif elem in second_grade_list:
                status_to_write["second_status"] = elem
                status_to_write["third_status"] = "None"
                break

            elif elem in third_grade_list:
                status_to_write["third_status"] = elem
                break

            dict_to_write = row.to_dict()
            list_to_sql.append(frame_of_tables['name'][loc])
            # print(list_to_sql)
            list_to_sql.append(str(datetime.now()))
            # print(list_to_sql)
            list_to_sql = list_to_sql + list(dict_to_write.values())
            # print(list_to_sql)
            list_to_sql = list_to_sql + list(status_to_write.values())
            # print(dict_to_write)
            cursor.executemany("""INSERT INTO out VALUES (?,?,?,?,?,?,?,?,?,?);""", (list_to_sql,))
            list_to_sql = []
            conn.commit()

    print("END")


# подключение или создание базы данных
def open_db():
    os.chdir("../second_output")
    db_name = "second_out.db"
    conn = sqlite3.connect(db_name)

    return conn


# создание таблицы в бд
def create_table_out(conn):
    cursor = conn.cursor()
    cursor.execute("""CREATE TABLE IF NOT EXISTS out(
                        filename VARCHAR,
                        created_dt DATETIME,
                        product_name VARCHAR,
                        article VARCHAR,
                        wholesale DECIMAL(9, 2),
                        retail_price DECIMAL(9, 2),
                        stock INTEGER,
                        first_status TEXT,
                        second_status TEXT,
                        third_status TEXT
                        );""")
    conn.commit()


if __name__ == "__main__":
    name = get_table()
    frame_of_tables = pars_table(name)
    conn = open_db()
    create_table_out(conn)
    make_table_for_sql(frame_of_tables, conn)
