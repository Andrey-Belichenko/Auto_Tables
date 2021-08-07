# -*- coding: utf-8 -*-
import os
import datetime
import logging as logger
import sqlite3
from sqlite3 import Error
import pandas as pd
import openpyxl as op
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font


def rename_table(name_of_table):
    """
    добавление текущей даты к названию таблицы
    :param name_of_table: имя таблицы
    :return: имя таблицы + текущая дата
    """
    name_of_table = name_of_table.replace(".xlsx", "")
    name_of_table = name_of_table + "_" + str(datetime.datetime.now().date())
    name_of_table = name_of_table + ".xlsx"
    return name_of_table


def create_logger(log_file_name, path='.'):
    """
    Создание и настройка логгера
    :param log_file_name: имя файла для записи логов
    :param path:
    """
    if not os.path.exists(path):
        os.mkdir(path)
    os.chdir(path)

    if not (log_file_name in os.listdir()):
        my_file = open(log_file_name, "w+")
        my_file.close()

    file_log = logger.FileHandler(log_file_name)
    console_out = logger.StreamHandler()

    logger.basicConfig(handlers=(file_log, console_out),
                       format='[%(asctime)s | %(levelname)s]: %(message)s',
                       datefmt='%m.%d.%Y %H:%M:%S',
                       level=logger.INFO)
    # возврат в корневую директорию проекта
    os.chdir("../")


def get_tables(file_mask, path='.'):
    """
    получение списка таблиц в path
    :param file_mask: маска поиска файлов
    :param path:
    :return: список имен файлов в виде list
    """
    logger.info("start reading input files")
    if not os.path.exists(path):
        logger.warning(msg=f'folder {path} does not exist')
        quit()

    os.chdir(path)
    tables_list = []
    for file in os.listdir():
        if file.find(file_mask) != -1:
            tables_list.append(file)

    if len(tables_list) == 0:
        logger.warning(msg=f'В каталоге {path} отсутствуют файлы "{file_mask}" для обработки.')
        exit(0)

    logger.info(tables_list)
    logger.info('tables list was generated')
    # возврат в корневую директорию проекта
    os.chdir("../")
    return tables_list


def parsing_tables(tables_list, path='.'):
    """
    Функция создания общей таблицы с данными из файлов input
    :param tables_list: список с названиями файлов входных таблиц
    :param path:
    :return: общаяя таблица в формате pandas dataframe
    """
    # список заголовков большой общей таблицы
    list_of_headers = ["Номер заказа", "Статус", "Время создания заказа", "Время оплаты", "Стоимость товаров, Руб",
                       "Стоимость доставки, Руб", "Сумма заказа, Руб", "Скидка магазина, Руб", "Оплачено клиентом, Руб",
                       "Сумма возврата, Руб", "Возвраты", "Товары", "Артикул", "Id товаров",
                       "Примечания к заказу (покупателя)", "Примечания к заказу (продавца)", "Имя получателя", "Страна",
                       "Штат/провинция", "Город", "Адрес", "Индекс", "Телефон", "Способ доставки", "Отгрузка истекает",
                       "Трекинг номер", "Время отправки", "Время подтверждения покупателем"]
    # template_of_dataframe переменная для инициализации переменной frame_of_tables в формате DataFrame
    template_of_dataframe = {header: list() for header in list_of_headers}
    # dictionary_to_write словарь используемый для построчной записи данных в DataFrame
    dictionary_to_write = {header: 0 for header in list_of_headers}

    frame_of_tables = pd.DataFrame(template_of_dataframe)
    os.chdir(path)
    logger.info("start generating a large shared table")

    start_row = 5  # Минимальное значения строк данного листа книги
    start_col = 1  # Минимальное значения строк данного листа книги

    # добавление новых значений в словарь и dataframe
    for table in tables_list:
        work_book = op.load_workbook(table)
        sheet = work_book["sheet1"]
        max_row = sheet.max_row  # Получение максимального значения строк данного листа книги
        max_col = sheet.max_column  # Получение максимального столбцов строк данного листа книги
        for row in sheet.iter_rows(min_row=start_row, min_col=start_col, max_row=max_row, max_col=max_col):
            index_of_headers = 0
            for cell in row:
                if cell.value == "":
                    dictionary_to_write[list_of_headers[index_of_headers]] = ""
                else:
                    dictionary_to_write[list_of_headers[index_of_headers]] = cell.value
                    if list_of_headers[index_of_headers] == 'Товары':
                        dictionary_to_write[list_of_headers[index_of_headers]] =\
                            dictionary_to_write[list_of_headers[index_of_headers]].\
                            replace("(Ships From: Russian Federation)", "")

                        dictionary_to_write[list_of_headers[index_of_headers]] = \
                            dictionary_to_write[list_of_headers[index_of_headers]]. \
                            replace("; Ships From: Russian Federation", "")
                index_of_headers += 1
                dictionary_to_write["Статус"] = table
                # ignore_index=True запись в dataframe без индекса
            frame_of_tables = frame_of_tables.append(dictionary_to_write, ignore_index=True)
    # frame_of_tables["Имя таблицы"] = list_of_tabs
    logger.info("large shared table was generated")
    # возврат в корневую директорию проекта
    os.chdir("../")
    return frame_of_tables


def generate_products_xlsx(frame_of_tables, name_of_table, path='.'):
    """
      создание и запись таблицы товары
      :param frame_of_tables: большая таблица со всеми данными
      :param name_of_table: имя таблицы для сохранение
      :param path:
      :return:
      """
    # Поля которые необходимо включить в таблицу Товары.xlsx
    # Товары|Артикул|Количество
    data_frame_goods = pd.DataFrame({'Товары': [], 'Количество': []})
    frame_with_articles = frame_of_tables[['Товары', 'Артикул']]
    frame_of_tables = frame_of_tables[['Товары']]
    df_to_save = pd.DataFrame({'Товары': [], 'Количество': [], 'Артикул': []})
    names = []
    nums = []
    dictionary_to_write = {'Товары': 0, 'Количество': 0}
    # = "Товары.xlsx"

    if not os.path.exists(path):
        os.mkdir(path)
    os.chdir(path)

    list_of_names_tabs = os.listdir()

    logger.info(list_of_names_tabs)

    if not (name_of_table in list_of_names_tabs):
        logger.info(msg=f'{name_of_table}not found in directory /output')
        logger.info(msg=f'File {name_of_table} creation')
        wb = op.Workbook()
        wb.save(filename=name_of_table)
        logger.info(msg=f'Table {name_of_table} was created')
    else:
        logger.info(msg=f'{name_of_table} already created')
        name_of_table = rename_table(name_of_table)
        logger.info(msg=f'{name_of_table} create table with now date and time in name')
        wb = op.Workbook()
        wb.save(filename=name_of_table)

    new_string = ""
    list_of_names = frame_of_tables['Товары'].values.tolist()

    # "Какая-то ацкая конструкция"
    # Для разделения товаров (если несколько записанно в одну ячейку)
    for i in list_of_names:
        new_string = new_string + i + "\n"

    list_of_names = new_string.split('\n')

    list_of_names = [str(name) for name in list_of_names if str(name) != '']

    # name[4:] для отделения цифры в []
    list_of_names = [name[4:] for name in list_of_names]

    # тут отделяем кол-во от имен товаров
    for name in list_of_names:
        last_index_of_name = name.index("- Количество:")
        name_of_position = name[:last_index_of_name]
        number = name[last_index_of_name:]
        names.append(name_of_position)
        number = number.replace('- Количество: ', '')
        number = number[:number.index(' ')]
        nums.append(number)

    for index in range(len(names)):
        dictionary_to_write['Товары'] = names[index]
        dictionary_to_write['Количество'] = int(nums[index])
        # ignore_index=True для того что бы запись дополнительно не индексировалаь (не добавлялся индекс в начало)
        data_frame_goods = data_frame_goods.append(dictionary_to_write, ignore_index=True)
    set_names = set(names)

    for index in range(len(set_names)):
        df = data_frame_goods[data_frame_goods['Товары'] == list(set_names)[index]]
        dictionary_to_write['Товары'] = list(set_names)[index]
        dictionary_to_write['Количество'] = df['Количество'].sum()
        # ignore_index=True для того что бы запись дополнительно не индексировалаь (не добавлялся индекс в начало)
        df_to_save = df_to_save.append(dictionary_to_write, ignore_index=True)

    logger.info("table goods was generated")

    # сортировка по алфавиту в столбце товары
    logger.info("table sorting ...")
    df_to_save = df_to_save[['Товары', 'Количество']].sort_values(by='Товары')
    # index=False для записи в excel без доп. индексов
    logger.info("table was sorted")

    # generate col articles
    logger.info("generate articles col")
    # сдесь генерируем столбец артиклов
    for index in range(len(df_to_save['Товары'])):
        name = df_to_save['Товары'][index]
        for second_index in range(len(frame_with_articles['Товары'])):
            # сверяем кол-во завершающих строку символов те сверяем количество имен товаров и артиклов
            # тем самым можем понять можно ли однозначно опредилить артикул
            if name in frame_with_articles['Товары'][second_index]:
                if frame_with_articles['Товары'][second_index].count('\n') == \
                        frame_with_articles['Артикул'][second_index].count('\n'):
                    list_of_names = frame_with_articles['Товары'][second_index].split("\n")
                    list_of_articles = frame_with_articles['Артикул'][second_index].split("\n")
                    for third_index in range(len(list_of_names)):
                        if name in list_of_names[third_index] and list_of_articles[third_index] != '':
                            df_to_save.loc[index, 'Артикул'] = list_of_articles[third_index].split(" * ")[0]
                        else:
                            df_to_save.loc[index, 'Артикул'] = None
                else:
                    df_to_save.loc[index, 'Артикул'] = None
    logger.info("col articles was generated successfully")

    df_to_save[['Товары', 'Количество', 'Артикул']].to_excel(name_of_table, index=False)

    logger.info(msg=f'data was saved to {name_of_table}')
    make_table_style_products_xlsx(name_of_table)
    # возврат в корневую директорию проекта
    os.chdir("../")


def generate_parcels_xlsx(frame_of_tables, name_of_table, path='.'):
    """
    создание и запись таблицы посылки
    :param frame_of_tables: большая таблица со всеми данными
    :param name_of_table: имя таблицы для сохранение
    :param path:
    :return:
    """
    # Поля таблицы Посылки.xlsx
    # Номер заказа | Статус | Время создания заказа | Время оплаты | Стоимость Товаров | Стоимость доставки
    # | Сумма заказа| Скидка магазина | Оплачено клиентом | Сумма возврата | Возвраты | Товары
    # | Артикул | Id товара | Примечяние заказа (покупателя) | Примечание заказа (продавца) | Имя получателя
    # | Штат/провинция | Страна | Город | Телефон | Номер трекинга

    status_dict = {"info@vorobey-club.ru": 1,
                   "sadovyy-ioj@mail.ru": 2,
                   "hunting-club-store@mail.ru": 4,
                   "vorobey-club1@bk.ru": 6,
                   "vorobey-club3@bk.ru": 7,
                   "ammunition-store@mail.ru": 8}

    list_of_headers = ["Номер заказа", "Статус", "Товары", "Имя получателя",
                       "Штат/провинция", "Город", "Телефон", "Трекинг номер"]

    output_list = ["Номер заказа", "Статус", "Товары", "Количество",
                   "Имя получателя", "Штат/провинция", "Город", "Телефон", "Номер трекинга"]

    index_list = []
    len_list = []
    count_list = []
    index_phone_list = []

    template_dict = {header: list() for header in list_of_headers}

    dictionary_to_write = {header: 0 for header in list_of_headers}
    data_frame_parcels = pd.DataFrame(template_dict)

    out_template_dict = {header: list() for header in output_list}
    temp_df = pd.DataFrame(out_template_dict)

    dict_to_out = {header: list() for header in output_list}

    if not os.path.exists(path):
        os.mkdir(path)
    os.chdir(path)

    list_of_names_tabs = os.listdir()
    logger.info(list_of_names_tabs)

    if not (name_of_table in list_of_names_tabs):
        logger.info(msg=f'{name_of_table}not found in directory /output')
        logger.info(msg=f'File {name_of_table} creation')
        wb = op.Workbook()
        wb.save(filename=name_of_table)
        logger.info(msg=f'Table {name_of_table} was created')
    else:
        logger.info(msg=f'{name_of_table} already created')
        name_of_table = rename_table(name_of_table)
        logger.info(msg=f'{name_of_table} create table with now date and time in name')
        wb = op.Workbook()
        wb.save(filename=name_of_table)

    shape = frame_of_tables.shape
    len_of_dataframe = shape[0]
    logger.info(msg=f'Table {name_of_table} generate...')

    for index in range(len_of_dataframe):
        for header_index in range(len(list_of_headers)):
            dictionary_to_write[list_of_headers[header_index]] = \
                frame_of_tables.loc[index, list_of_headers[header_index]]
        # ignore_index=True для того что бы запись дополнительно не индексировалаь (не добавлялся индекс в начало)
        data_frame_parcels = data_frame_parcels.append(dictionary_to_write, ignore_index=True)

    for index in range(len(frame_of_tables.index)):
        obj = frame_of_tables["Товары"][index]
        if obj.count('\n') > 0:
            obj_list = obj.split('\n')
            len_list.append(index + len(obj_list))
            for one in obj_list:
                name = one[4:one.find("- Количество: ")]
                # print(name)
                num = one[one.find("Количество: ") + len("Количество: "):]
                num = num[:num.find(" ")]
                status = status_dict[(frame_of_tables["Статус"][index].split())[0]]
                # print(num)
                dict_to_out["Товары"].append(name)
                dict_to_out["Количество"].append(num)
                dict_to_out["Номер заказа"].append(frame_of_tables["Номер заказа"][index])
                dict_to_out["Статус"].append(status)
                dict_to_out["Имя получателя"].append(frame_of_tables["Имя получателя"][index])
                dict_to_out["Штат/провинция"].append(frame_of_tables["Штат/провинция"][index])
                dict_to_out["Город"].append(frame_of_tables["Город"][index])
                dict_to_out["Телефон"].append(frame_of_tables["Телефон"][index])
                dict_to_out["Номер трекинга"].append(frame_of_tables["Трекинг номер"][index])

        else:
            name = obj[4:obj.find("- Количество: ")]
            num = obj[obj.find("Количество: ") + len("Количество: "):]
            num = num[:num.find(" ")]
            status = status_dict[(frame_of_tables["Статус"][index].split())[0]]

            dict_to_out["Товары"].append(name)
            dict_to_out["Количество"].append(num)
            dict_to_out["Статус"].append(status)
            dict_to_out["Номер заказа"].append(frame_of_tables["Номер заказа"][index])
            dict_to_out["Имя получателя"].append(frame_of_tables["Имя получателя"][index])
            dict_to_out["Штат/провинция"].append(frame_of_tables["Штат/провинция"][index])
            dict_to_out["Город"].append(frame_of_tables["Город"][index])
            dict_to_out["Телефон"].append(frame_of_tables["Телефон"][index])
            dict_to_out["Номер трекинга"].append(frame_of_tables["Трекинг номер"][index])

    out_df = pd.DataFrame(dict_to_out)
    old_order_id = ""
    old_flag = 0

    # востановление порядка индексов после вставок датафреймов
    out_df = out_df.reset_index(drop=True)

    # далее премещаем записи с одинаковыми номерами телефонов друг к другу
    list_of_index = pars_coll_numbers(out_df)
    dictionary_to_write = {header: 0 for header in output_list}
    logger.info("make dataframe by phone numbers")
    for index in list_of_index:
        dictionary_to_write["Номер заказа"] = out_df["Номер заказа"][index]
        dictionary_to_write["Статус"] = out_df["Статус"][index]
        dictionary_to_write["Товары"] = out_df["Товары"][index]
        dictionary_to_write["Количество"] = out_df["Количество"][index]
        dictionary_to_write["Имя получателя"] = out_df["Имя получателя"][index]
        dictionary_to_write["Штат/провинция"] = out_df["Штат/провинция"][index]
        dictionary_to_write["Город"] = out_df["Город"][index]
        dictionary_to_write["Телефон"] = out_df["Телефон"][index]
        dictionary_to_write["Номер трекинга"] = out_df["Номер трекинга"][index]
        temp_df = temp_df.append(dictionary_to_write, ignore_index=True)
    out_df = temp_df

    # генерация списка номеров строк для обединениия
    logger.info("generate list of index to marge")
    for index in range(len(out_df["Номер заказа"])):
        order_id = out_df["Номер заказа"][index]
        # print(out_df["Имя получателя"][index], index)
        if order_id == old_order_id and old_flag == 0:
            index_list.append(index + 1)
            old_flag = 1
        if order_id != old_order_id and old_flag == 1:
            index_list.append(index + 1)
            old_flag = 0
        old_order_id = order_id

    # собираем индексы ячеек столбца трек-номер который следует выделить жирным
    old_phone = ""
    for index in range(len(out_df["Телефон"])):
        phone = out_df["Телефон"][index]
        if phone == old_phone and old_flag == 0:
            index_phone_list.append(index + 1)
            old_flag = 1
        if phone != old_phone and old_flag == 1:
            index_phone_list.append(index + 1)
            old_flag = 0
        old_phone = phone
    out_df = out_df.reset_index(drop=True)

    # собираем список индексов которые надо выделить жирным
    logger.info("generate list of index to count")
    for index in range(len(out_df['Количество'])):
        if int(out_df['Количество'][index]) > 1:
            count_list.append(index)

    logger.info(msg=f'Save table as {name_of_table}')
    # index=False для записи в excel без доп. индексов
    out_df.to_excel(name_of_table, index=False)

    logger.info(msg=f'data was saved to {name_of_table}')

    make_table_style_parcels_xlsx(name_of_table, count_list, index_phone_list)

    make_merge(index_list, name_of_table)

    os.chdir("../")


def generate_tracking_xlsx(frame_of_tables, list_of_tables_loc, path='.'):
    """
    Создание таблиц с трек-номерами для работы с aliexpress
    :param frame_of_tables: большой dataframe со всеми данными
    :param list_of_tables_loc: список имен файлов содержаших таблицы
    :param path:
    """

    list_of_headers_eng = ['Order Number',  'Logistics Company', 'Tracking Number']

    const_value_dict = {'Статус': "Full Shipment", 'Заметка': "",
                        'Сайт трекинга': "http://www.russianpost.ru/tracking20/"}

    template_dict = {header: list() for header in list_of_headers_eng}

    df_to_write = pd.DataFrame(template_dict)

    # дополнительная чать имени файла с трекинг таблицами
    file_name_add = "- butch send.xlsx"

    if not os.path.exists(path):
        os.mkdir(path)

    os.chdir(path)

    for name in list_of_tables_loc:
        name_of_table = name.replace(".xlsx", "")+file_name_add
        if not os.path.exists(name_of_table):
            logger.info(msg=f'{name_of_table}not found in directory /output')
            logger.info(msg=f'File {name_of_table} creation')
            wb = op.Workbook()
            wb.save(filename=name_of_table)
            logger.info(msg=f'Table {name_of_table} was created')
        else:
            logger.info(msg=f'{name_of_table} already created')
            name_of_table = rename_table(name_of_table)
            logger.info(msg=f'{name_of_table} create table with now date and time in name')
            wb = op.Workbook()
            wb.save(filename=name_of_table)
        # выбираем из общей кучи только записи из конкретной таблицы
        frame_of_tables_mod = frame_of_tables[frame_of_tables["Статус"].isin([name])]
        frame_of_tables_mod = frame_of_tables_mod[["Номер заказа", "Способ доставки", "Трекинг номер"]]
        logger.info(f"Create {name_of_table}")
        for index in range(len(frame_of_tables_mod)):
            # Заполняем таблицу с трекинга даннмыи
            dict_to_write = \
                dict(zip(list_of_headers_eng,
                         frame_of_tables_mod.iloc[index]))
            dict_to_write['Delivery Status'] = const_value_dict['Статус']
            dict_to_write['Remark'] = const_value_dict['Заметка']
            dict_to_write['Tracking website'] = const_value_dict['Сайт трекинга']
            df_to_write = df_to_write.append(dict_to_write, ignore_index=True)
        df_to_write.to_excel(name_of_table, index=False)
        df_to_write = df_to_write.iloc[0:0]
        make_table_style_tracking_xlsx(name_of_table)

    os.chdir('../')


def create_connection(path):
    """
     подключение к базе данных SQLite
    :param path:
    """
    try:
        connection = sqlite3.connect(path)
        logger.info("Connection to SQLite DB successful")
        return connection

    except Error as e:
        # прграмма завершиться с ошибкой если не подключится
        logger.error(msg=f'The error {e} occurred')
        quit()


def execute_query(connection, query):
    """
    исполнение запроса SQLite и сохранение изменений
    :param connection: ваше подключение к БД
    :param query: команда БД для исполнения
    """
    cursor = connection.cursor()
    cursor.execute(query)
    connection.commit()


def generate_sqlite(name, path='.'):
    """
    создание юазы данных SQLite
    :param name: имя БД
    :param path:
    :return: connection, name
    """
    if not os.path.exists(path):
        os.mkdir(path)
    os.chdir(path)
    connection = create_connection(name)
    os.chdir('../')
    return connection, name


def dataframe_to_sqlite(frame_of_tables, conn):
    """
     запись данных в таблицу БД SQLite
    :param frame_of_tables:  большая сводная таблица pandas dataframe
    :param conn: подключение SQLite
    """
    list_of_headers = ["order_id", "status", "order_credit_dt", "pay_dt", "goods_cost", "delivery_cost",
                       "order_cost", "shop_discount", "customer_paid", "refund", "returns", "products",
                       "products_articles", "product_ids", "buyer_note", "seller_note", "delivery_name",
                       "delivery_country", "delivery_state", "delivery_city", "delivery_address", "delivery_zip",
                       "delivery_phone", "delivery_type", "delivery_send_until_date", "delivery_tracking_number",
                       "delivery_send_dt", "delivery_receive_dt"]

    rus_headers = ["Номер заказа", "Статус", "Время создания заказа", "Время оплаты", "Стоимость товаров, Руб",
                   "Стоимость доставки, Руб", "Сумма заказа, Руб", "Скидка магазина, Руб", "Оплачено клиентом, Руб",
                   "Сумма возврата, Руб", "Возвраты", "Товары", "Артикул", "Id товаров",
                   "Примечания к заказу (покупателя)", "Примечания к заказу (продавца)", "Имя получателя", "Страна",
                   "Штат/провинция", "Город", "Адрес", "Индекс", "Телефон", "Способ доставки", "Отгрузка истекает",
                   "Трекинг номер", "Время отправки", "Время подтверждения покупателем"]

    logger.info('write to database...')

    cursor = conn.cursor()
    frame_of_tables = frame_of_tables[rus_headers]
    for index in range(len(list_of_headers)):
        frame_of_tables = frame_of_tables.rename(columns={rus_headers[index]: list_of_headers[index]})

    # frame_of_tables.to_sql(db_name, con=connection, schema='dbo', if_exists='replace')

    for loc in range(len(frame_of_tables.index)):
        row = frame_of_tables.iloc[loc]
        list_to_write = list(row)
        cursor.executemany("""INSERT INTO orders VALUES
                            (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?);""", (list_to_write,))
        conn.commit()

    logger.info(msg=f'database {db_name} successfully')


def create_sqlite_table(conn):
    """
    создание таблицы в БД SQLite с использованием запроса SQLite
    :param conn: подключение SQLite
    """
    cursor = conn.cursor()
    cursor.execute("""CREATE TABLE IF NOT EXISTS orders(
                            order_id VARCHAR(16), 
                            status VARCHAR(255), 
                            order_created_dt DATETIME, 
                            pay_dt DATETIME, 
                            goods_cost DECIMAL(9,2), 
                            delivery_cost DECIMAL(9,2), 
                            order_cost DECIMAL(9,2), 
                            shop_discount DECIMAL(9,2), 
                            refund DECIMAL(9,2), 
                            client paid DECIMAL(9,2),
                            returns VARCHAR(1024), 
                            products TEXT, 
                            product_articles TEXT, 
                            product_ids TEXT, 
                            buyer_note TEXT, 
                            seller_note TEXT, 
                            delivery_name TEXT, 
                            delivery_country TEXT, 
                            delivery_state TEXT, 
                            delivery_city TEXT, 
                            delivery_address TEXT, 
                            delivery_zip VARCHAR(50), 
                            delivery_phone VARCHAR(50), 
                            delivery_type VARCHAR(255), 
                            delivery_send_until_date DATETIME, 
                            delivery_tracking_number VARCHAR(255), 
                            delivery_send_dt DATETIME, 
                            delivery_receive_dt DATETIME
                            );""")
    conn.commit()


def make_table_style_products_xlsx(name):
    """
    приведение таблицы товары к соответствующему виду
    :param name: имя xlsx файла
    """
    work_book = op.load_workbook(name)
    col_letters = ['A', 'B', 'C']  # список букв колонок
    sheet = work_book.active
    for letter in col_letters:
        if letter == 'A':
            # изменение щирины колонки A
            sheet.column_dimensions[letter].width = 130
    work_book.save(name)


def make_table_style_parcels_xlsx(name, count_list, index_phone_list):
    """
    приведение таблицы посылки к соответствующему виду
    :param name: имя xlsx файла
    :param count_list: список индексов для выделения жирным количества больше 1
    :param index_phone_list: список индексов для выделения трек-номера посылок с одинаковыми номерами телефонов
    """
    work_book = op.load_workbook(name)
    # col_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']  # список букв колонок
    col_letters = {'A': 16.44, 'B': 1, 'C': 57, 'D': 2.56, 'E': 15.89, 'F': 15.89, 'G': 13.11, 'H': 14.33, 'I': 14.33}
    thin = Side(border_style="thin", color="000000")

    sheet = work_book.active
    for letter in col_letters.keys():
        # изменение ширины столбца взависимости от его буквы
        sheet.column_dimensions[letter].width = col_letters[letter] + 0.8

        for cell in sheet[letter]:
            # задача общих параметров для всех ячеек
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for index in range(sheet.max_row):
        # настройка ширины для строк
        sheet.row_dimensions[index].height = 43.2

    # сдесь выделяем первую строку для настройки особых параметров
    first_row = sheet[1]
    sheet.row_dimensions[1].height = 15
    for cell in first_row:
        # настройка заливки цвет и тип
        cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        # настройка шрифта жирность и цвет
        cell.font = Font(color="000000", bold=True)

    # тут выделяем жирным количество больше 1
    logger.info("make count bold")
    coll_d = sheet['D']
    for index in count_list:
        # index + 1 для преревода в отчет от 1 а не от 0
        coll_d[index + 1].fill = PatternFill(start_color="D8D8D8", end_color="D8D8D8", fill_type="solid")
        # index + 2 тк из-за наложения форматоыв индекс чутьчуть плывет
        sheet['D' + str(index + 2)].font = Font(bold=True)

    # тут выделяем жирнвм трек-номера посылок с одинаковыми номерами телефонов
    logger.info("make track-number bold")
    for index in range(0, len(index_phone_list), 2):
        if len(range(index_phone_list[index], index_phone_list[index + 1])) == 1:
            sheet['I' + str(index_phone_list[index])].font = Font(bold=True)
            sheet['I' + str(index_phone_list[index + 1])].font = Font(bold=True)
        else:
            for sub_index in range(index_phone_list[index], index_phone_list[index + 1]):
                sheet['I' + str(sub_index + 1)].font = Font(bold=True)

    work_book.save(name)


def make_table_style_tracking_xlsx(name):
    """
    Задание необходимых стилей таблицам с трек-номерами
    :param name: имя таблицы
    """
    work_book = op.load_workbook(name)

    col_let = ['A', 'B', 'C', 'D', 'E', 'F']

    sheet = work_book.active

    # настройка ширины столбцов в таблице
    logger.info(f"make styles in {name}")
    for letter in col_let:
        sheet.column_dimensions[letter].width = 31.1

    # задача параметров текста в ячейках
    for letter in col_let:
        for cell in sheet[letter]:
            if letter == 'A' or letter == 'D':
                cell.font = Font(name="Colibri", size=11)
            else:
                cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # сдесь выделяем первую строку для настройки особых параметров
    first_row = sheet[1]
    for cell in first_row:
        # настройка заливки цвет и тип
        cell.fill = PatternFill(start_color="00CDFF", end_color="00CDFF", fill_type="solid")
        # настройка шрифта жирность
        cell.font = Font(name="Arial", size=12, color="810081", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

    work_book.save(name)


def make_merge(index_list, name):
    """
    создание объединенных ячеек таблице
    :param index_list: список индексов для объединения ячеек
    :param name: имя xlsx файла
    :return:
    """
    # Буквы столбцов в которых одинаковые данные должны объединяться
    merge_letters = ['A', 'B', 'E', 'F', 'G', 'H', 'I']
    work_book = op.load_workbook(name)
    sheet = work_book.active
    logger.info("make cells merge")
    for letter in merge_letters:
        # выбираем из списка индексы с шагом 2 тк каждый 2 закрывающий
        for index in range(0, len(index_list), 2):
            # строка обединениея передаем в sheet.merge_cells строку вида "БУКВА+ЦИФРА:БУКВА+ЦИФРА" ("A1:A4")
            sheet.merge_cells(letter + str(index_list[index]) + ':' + letter + str(index_list[index + 1]))
    work_book.save(name)


def pars_coll_numbers(out_df):
    """
    сортировка списка по наличию одинаковых номеров телефонов
    :param out_df: выходной dataframe
    :return: список индексов под сдвиг
    """
    phone_list = out_df['Телефон']
    phone_list = phone_list.to_list()
    list_of_index = []
    temp = []
    # создаем список телефонов без повторений
    for x in phone_list:
        if x not in temp:
            temp.append(x)
    phone_set = temp
    # создание списка индексов под сдвиг
    for phone_number in phone_set:
        for index in range(len(phone_list)):
            if phone_number == phone_list[index]:
                list_of_index.append(index)
    return list_of_index


def move_old_files(list_of_tables_loc, archive_path='Archive', input_path='input'):
    """
    Функция перемещает старые входные файлы в архив
    :param list_of_tables_loc : список имен xlsx файлов
    :param archive_path: имя папки архива
    :param input_path: имя папки с входнами данными
    """
    if not os.path.exists(archive_path):
        os.makedirs(archive_path)
        logger.info(f"folder {archive_path} was created")

    logger.info("moving old files to archive folder")
    for name in list_of_tables_loc:
        logger.info(input_path+"/"+name+" -> "+archive_path+"/"+name)
        os.rename(input_path+"/"+name, archive_path+"/"+name)


if __name__ == '__main__':
    create_logger("logs.log", "logger")
    list_of_tables = get_tables(".xlsx", "input")
    frame_of_tables_g = parsing_tables(list_of_tables, "input")
    generate_products_xlsx(frame_of_tables_g, "Товары.xlsx", "output")
    generate_parcels_xlsx(frame_of_tables_g, "Посылки.xlsx", "output")
    generate_tracking_xlsx(frame_of_tables_g, list_of_tables, "output")
    connect, db_name = generate_sqlite("out.db", "output")
    create_sqlite_table(connect)
    dataframe_to_sqlite(frame_of_tables_g, connect)
    move_old_files(list_of_tables, 'Archive', 'input')
    logger.info('the program ended successfully')
